import os
import cv2
import time
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
from datetime import datetime
from openpyxl import Workbook, load_workbook
from takeImage import capture_images_for_enrollment
from trainimage import TrainImage

# ----------- PATH SETUP -----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
haarcasecade_path = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")
trainimage_path = os.path.join(BASE_DIR, "TrainingImage")
trainimagelabel_path = os.path.join(BASE_DIR, "TrainingImageLabel", "Trainner.yml")
attendance_excel = os.path.join(BASE_DIR, "attendance.xlsx")
labels_file = os.path.join(BASE_DIR, "labels.txt")

# ----------- EXCEL ATTENDANCE ----------
def mark_attendance(name, subject):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    if not os.path.exists(attendance_excel):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Subject", "Date", "Time"])
    else:
        wb = load_workbook(attendance_excel)
        ws = wb.active

    ws.append([name, subject, date_str, time_str])
    wb.save(attendance_excel)

# ----------- SUBJECT DETECTION ----------
def get_subject():
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    if "07:00" <= current_time <= "08:00":
        return "DBMS"
    elif "08:10" <= current_time <= "09:10":
        return "Python"
    elif "09:20" <= current_time <= "10:20":
        return "C++"
    elif "10:30" <= current_time <= "11:30":
        return "C"
    elif "11:40" <= current_time <= "12:40":
        return "DSA"
    else:
        return "General"

# ----------- LOAD LABELS -----------
def load_labels():
    labels = {}
    try:
        with open(labels_file, "r") as f:
            for line in f:
                student_id, name = line.strip().split(",")
                labels[int(student_id)] = name
    except FileNotFoundError:
        print("[ERROR] labels.txt not found.")
    return labels

# ----------- FACE RECOGNITION ----------
def start_attendance():
    labels = load_labels()
    if not os.path.exists(trainimagelabel_path):
        messagebox.showerror("Error", "No trained model found. Register students first.")
        return

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read(trainimagelabel_path)

    face_cascade = cv2.CascadeClassifier(haarcasecade_path)
    cam = cv2.VideoCapture(0)

    messagebox.showinfo("Attendance", "Look at the camera to mark attendance.")
    timeout = 15
    start_time = time.time()
    found = False

    while True:
        ret, frame = cam.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            id_, conf = recognizer.predict(gray[y:y+h, x:x+w])
            if conf < 60:
                name = labels.get(id_)
                if name:
                    mark_attendance(name, get_subject())
                    cv2.putText(frame, f"{name} - Present", (x, y-10),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                    found = True
                    break
                else:
                    cv2.putText(frame, "Unknown", (x, y-10),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
            else:
                cv2.putText(frame, "Unknown", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        cv2.imshow("Attendance", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        if found or time.time() - start_time > timeout:
            break

    cam.release()
    cv2.destroyAllWindows()

# ----------- REGISTRATION UI ----------
def open_register_ui():
    reg = Toplevel(window)
    reg.title("Register Student")
    reg.geometry("600x360")
    Label(reg, text="Enrollment:", font=("Verdana",12)).place(x=50,y=60)
    ent_en = Entry(reg, font=("Verdana",14)); ent_en.place(x=180,y=60)
    Label(reg, text="Name:", font=("Verdana",12)).place(x=50,y=110)
    ent_name = Entry(reg, font=("Verdana",14)); ent_name.place(x=180,y=110)
    Label(reg, text="Subject:", font=("Verdana",12)).place(x=50,y=160)
    ent_sub = Entry(reg, font=("Verdana",14)); ent_sub.place(x=180,y=160)
    Label(reg, text="Start (HH:MM):", font=("Verdana",12)).place(x=50,y=210)
    ent_start = Entry(reg, font=("Verdana",14)); ent_start.place(x=180,y=210)
    Label(reg, text="End (HH:MM):", font=("Verdana",12)).place(x=50,y=260)
    ent_end = Entry(reg, font=("Verdana",14)); ent_end.place(x=180,y=260)

    def do_register():
        enrollment = ent_en.get().strip()
        name = ent_name.get().strip()
        subject = ent_sub.get().strip() or "General"
        start = ent_start.get().strip() or "00:00"
        end = ent_end.get().strip() or "23:59"

        if not enrollment or not name:
            messagebox.showwarning("Input", "Enrollment and Name required")
            return

        captured = capture_images_for_enrollment(enrollment, name, haarcasecade_path,
                                                trainimage_path, subject, start, end)
        if captured:
            messagebox.showinfo("Captured", "Images captured successfully. Training model now...")
            trained = TrainImage(haarcasecade_path, trainimage_path, trainimagelabel_path, None)
            if trained:
                messagebox.showinfo("Done", "Model trained successfully.")
            else:
                messagebox.showwarning("Train", "Training failed - please check images.")
            reg.destroy()
        else:
            messagebox.showerror("Failed", "Image capture failed. Try again.")

    Button(reg, text="Register & Train", command=do_register,
           font=("Verdana",12), bg="black", fg="yellow").place(x=240,y=300)

# ----------- MAIN WINDOW ----------
window = Tk()
window.title("Face Recognition Attendance System")
window.geometry("1280x720")
window.configure(background="#1c1c1c")

Button(window, text="Register a New Student", command=open_register_ui,
       font=("Verdana",16), bg="black", fg="yellow").pack(pady=20)

Button(window, text="Start Attendance", command=start_attendance,
       font=("Verdana",16), bg="black", fg="lightgreen").pack(pady=20)

window.mainloop()
